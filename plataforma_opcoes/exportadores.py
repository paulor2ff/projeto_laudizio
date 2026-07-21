"""
exportadores.py — geração de exportações (cotações/opções em csv/xlsx/pdf)
como bytes em memória.

Usado por dois consumidores diferentes:
  - cli.py    grava os bytes num ficheiro (--exportar --formato ...)
  - api.py    devolve os bytes como resposta HTTP (GET /exportar/...)

A mesma lógica de montagem, duas formas de entregar o resultado — evita
ter o layout do .xlsx/.pdf definido em dois sítios que podem divergir.

Cada função `*_bytes()` devolve uma tupla (conteudo, contagem):
  - conteudo:  bytes prontos para gravar/enviar, ou None se não há dados
  - contagem:  nº de registros/contratos (para mensagens de status)
"""

import csv
import io
from datetime import date, datetime

from database import consultar_cotacoes, consultar_opcoes

MEDIA_TYPES = {
    "csv":  "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf":  "application/pdf",
}


def _nome_exportacao(prefixo: str, ticker: str, extensao: str) -> str:
    """Nome padrão de arquivo de exportação: prefixo_TICKER_AAAA-MM-DD.ext"""
    return f"{prefixo}_{ticker.replace('.SA', '')}_{date.today()}.{extensao}"


# ─── Formatação — usada pelas próprias tabelas .pdf e por cli.py ────────────

def _fmt(v, casas=2, default="—"):
    if v is None:
        return default
    try:
        return f"{float(v):.{casas}f}"
    except Exception:
        return str(v)


def _fmt_i(v, default="—"):
    if v is None:
        return default
    try:
        return f"{int(v):,}".replace(",", ".")
    except Exception:
        return str(v)


def _fmt_pct(v, default="—"):
    if v is None:
        return default
    try:
        return f"{float(v):+.2f}%"
    except Exception:
        return str(v)


def _fmt_m(v, default="—"):
    """Formata valor monetário em K/M."""
    if v is None:
        return default
    try:
        v = float(v)
        if v >= 1_000_000:
            return f"R$ {v/1_000_000:.1f}M"
        if v >= 1_000:
            return f"R$ {v/1_000:.1f}K"
        return f"R$ {v:.2f}"
    except Exception:
        return str(v)


# ─── CSV — dados "crus", iguais aos usados no banco ────────────────────────

def cotacoes_csv_bytes(ticker, data_inicio=None, data_fim=None):
    rows = consultar_cotacoes(ticker, data_inicio, data_fim)
    if not rows:
        return None, 0
    campos = ["ticker", "data", "abertura", "maxima", "minima", "fechamento", "adj_close", "volume"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=campos, extrasaction="ignore")
    w.writeheader()
    w.writerows([dict(r) for r in rows])
    return ("\ufeff" + buf.getvalue()).encode("utf-8"), len(rows)


def opcoes_csv_bytes(ticker, tipo=None, vencimento=None, data_inicio=None, data_fim=None):
    rows = consultar_opcoes(ticker, tipo, vencimento, data_inicio=data_inicio, data_fim=data_fim)
    if not rows:
        return None, 0
    campos = [
        "ticker_ativo","codigo","tipo","modelo","strike","vencimento",
        "ultimo","variacao_pct","data_hora","num_negocios","vol_financeiro",
        "vol_implícita","iq","coberto","descoberto","travado","titulares","lancadores",
        "delta","gamma","theta","vega","rho","otm_atm_itm","dist_strike","iq_calc",
    ]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=campos, extrasaction="ignore")
    w.writeheader()
    w.writerows([dict(r) for r in rows])
    return ("\ufeff" + buf.getvalue()).encode("utf-8"), len(rows)


# ─── .xlsx — mesmos dados do CSV, formatados para leitura humana ───────────
# (cabeçalho colorido, zebra, formato numérico, verde/vermelho para alta/baixa)

def cotacoes_xlsx_bytes(ticker, data_inicio=None, data_fim=None):
    rows = consultar_cotacoes(ticker, data_inicio, data_fim)
    if not rows:
        return None, 0

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ticker_limpo = ticker.replace(".SA", "")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotações"

    azul_titulo = PatternFill("solid", fgColor="1F4E78")
    azul_cab    = PatternFill("solid", fgColor="2E75B6")
    cinza_zebra = PatternFill("solid", fgColor="F2F2F2")
    branco_bold = Font(bold=True, color="FFFFFF")
    verde       = Font(color="1A7A3C")
    vermelho    = Font(color="B3261E")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Cotações Históricas — {ticker_limpo}"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = azul_titulo
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}  ·  {len(rows):,} registros"
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    cabecalhos = ["Ticker", "Data", "Abertura", "Máxima", "Mínima",
                  "Fechamento", "Fech. Ajustado", "Volume"]
    campos     = ["ticker", "data", "abertura", "maxima", "minima",
                  "fechamento", "adj_close", "volume"]
    linha_cab = 4
    for col, titulo in enumerate(cabecalhos, start=1):
        cel = ws.cell(row=linha_cab, column=col, value=titulo)
        cel.font = branco_bold
        cel.fill = azul_cab
        cel.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{linha_cab + 1}"

    for i, r in enumerate(rows):
        d = dict(r)
        linha = linha_cab + 1 + i
        for col, campo in enumerate(campos, start=1):
            cel = ws.cell(row=linha, column=col, value=d.get(campo))
            if campo == "data" and isinstance(d.get(campo), (date, datetime)):
                cel.number_format = "yyyy-mm-dd"
            elif campo in ("abertura", "maxima", "minima", "fechamento", "adj_close"):
                cel.number_format = "#,##0.00"
            elif campo == "volume":
                cel.number_format = "#,##0"
        if i % 2 == 1:
            for col in range(1, len(campos) + 1):
                ws.cell(row=linha, column=col).fill = cinza_zebra
        abertura, fechamento = d.get("abertura"), d.get("fechamento")
        if abertura is not None and fechamento is not None:
            cel_fech = ws.cell(row=linha, column=campos.index("fechamento") + 1)
            cel_fech.font = verde if fechamento >= abertura else vermelho

    for col, w in enumerate([12, 12, 11, 11, 11, 12, 14, 13], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows)


def opcoes_xlsx_bytes(ticker, tipo=None, vencimento=None, data_inicio=None, data_fim=None):
    rows = consultar_opcoes(ticker, tipo, vencimento, data_inicio=data_inicio, data_fim=data_fim)
    if not rows:
        return None, 0

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ticker_limpo = ticker.replace(".SA", "")
    wb = Workbook()
    ws = wb.active
    ws.title = "Opções"

    azul_titulo = PatternFill("solid", fgColor="1F4E78")
    azul_cab    = PatternFill("solid", fgColor="2E75B6")
    cinza_zebra = PatternFill("solid", fgColor="F2F2F2")
    branco_bold = Font(bold=True, color="FFFFFF")
    verde       = Font(color="1A7A3C")
    vermelho    = Font(color="B3261E")

    cabecalhos = ["Código","Tipo","Modelo","Strike","Vencimento","Último","Variação %",
                  "Data/Hora","Nº Negócios","Vol. Financeiro","Vol. Implícita %","IQ",
                  "Coberto","Descoberto","Travado","Titulares","Lançadores",
                  "Delta","Gamma","Theta","Vega","Rho","OTM/ATM/ITM","Dist. Strike","IQ Calc"]
    campos     = ["codigo","tipo","modelo","strike","vencimento","ultimo","variacao_pct",
                  "data_hora","num_negocios","vol_financeiro","vol_implícita","iq",
                  "coberto","descoberto","travado","titulares","lancadores",
                  "delta","gamma","theta","vega","rho","otm_atm_itm","dist_strike","iq_calc"]
    n_col = len(cabecalhos)
    ultima_col = get_column_letter(n_col)

    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"] = f"Cadeia de Opções — {ticker_limpo}"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = azul_titulo
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{ultima_col}2")
    ws["A2"] = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}  ·  {len(rows):,} contratos"
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    linha_cab = 4
    for col, titulo in enumerate(cabecalhos, start=1):
        cel = ws.cell(row=linha_cab, column=col, value=titulo)
        cel.font = branco_bold
        cel.fill = azul_cab
        cel.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{linha_cab + 1}"

    colunas_pct    = {"variacao_pct", "vol_implícita"}
    colunas_moeda  = {"strike", "ultimo", "vol_financeiro"}
    colunas_greek  = {"delta", "gamma", "theta", "vega", "rho"}

    for i, r in enumerate(rows):
        d = dict(r)
        linha = linha_cab + 1 + i
        for col, campo in enumerate(campos, start=1):
            cel = ws.cell(row=linha, column=col, value=d.get(campo))
            if campo == "vencimento" and isinstance(d.get(campo), (date, datetime)):
                cel.number_format = "yyyy-mm-dd"
            elif campo in colunas_moeda:
                cel.number_format = "#,##0.00"
            elif campo in colunas_pct:
                cel.number_format = "0.00"
            elif campo in colunas_greek:
                cel.number_format = "0.0000"
            elif campo in ("dist_strike", "iq", "iq_calc"):
                cel.number_format = "0.00"
            elif campo == "num_negocios":
                cel.number_format = "#,##0"
        if i % 2 == 1:
            for col in range(1, n_col + 1):
                ws.cell(row=linha, column=col).fill = cinza_zebra
        variacao = d.get("variacao_pct")
        if variacao is not None:
            cel_var = ws.cell(row=linha, column=campos.index("variacao_pct") + 1)
            cel_var.font = verde if variacao >= 0 else vermelho

    larguras = [12,6,10,9,11,9,10,16,10,14,13,7,8,10,8,9,10,8,8,8,8,8,11,10,8]
    for col, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows)


# ─── .pdf — visão curada e formatada, pensada para imprimir/ler ────────────
# (mesmas colunas já usadas em imprimir_cotacoes()/imprimir_opcoes() no terminal,
#  não as 25 colunas completas do CSV — um PDF é uma página fixa, não uma planilha)

def _tabela_pdf(destino, titulo, subtitulo, cabecalho, dados_linhas,
                 col_cor=None, fonte_tamanho=8):
    """destino: caminho (str) ou objeto tipo-arquivo (ex.: io.BytesIO)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    estilos = getSampleStyleSheet()
    doc = SimpleDocTemplate(destino, pagesize=landscape(A4),
                             leftMargin=1.3 * cm, rightMargin=1.3 * cm,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             title=titulo)
    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(subtitulo, estilos["Normal"]),
        Spacer(1, 0.6 * cm),
    ]

    dados = [cabecalho] + [linha for linha, _cor in dados_linhas]
    tabela = Table(dados, repeatRows=1)
    estilo = [
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), fonte_tamanho),
        ("ALIGN",          (1, 0), (-1, -1), "RIGHT"),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
    ]
    if col_cor is not None:
        for i, (_linha, cor) in enumerate(dados_linhas, start=1):
            if cor is not None:
                estilo.append(("TEXTCOLOR", (col_cor, i), (col_cor, i), cor))
    tabela.setStyle(TableStyle(estilo))
    elementos.append(tabela)
    doc.build(elementos)


def cotacoes_pdf_bytes(ticker, data_inicio=None, data_fim=None):
    rows = consultar_cotacoes(ticker, data_inicio, data_fim)
    if not rows:
        return None, 0
    from reportlab.lib import colors

    ticker_limpo = ticker.replace(".SA", "")
    cabecalho = ["Data", "Abertura", "Máxima", "Mínima", "Fechamento", "Volume"]
    dados_linhas = []
    for r in rows:
        d = dict(r)
        linha = [
            str(d.get("data"))[:10] if d.get("data") else "—",
            _fmt(d.get("abertura")), _fmt(d.get("maxima")),
            _fmt(d.get("minima")),   _fmt(d.get("fechamento")),
            _fmt_i(d.get("volume")),
        ]
        abertura, fechamento = d.get("abertura"), d.get("fechamento")
        cor = None
        if abertura is not None and fechamento is not None:
            cor = colors.HexColor("#1A7A3C") if fechamento >= abertura else colors.HexColor("#B3261E")
        dados_linhas.append((linha, cor))

    buf = io.BytesIO()
    _tabela_pdf(
        buf,
        titulo=f"Cotações Históricas — {ticker_limpo}",
        subtitulo=f"Gerado em {datetime.now():%d/%m/%Y %H:%M}  ·  {len(rows):,} registros",
        cabecalho=cabecalho, dados_linhas=dados_linhas, col_cor=4,
    )
    return buf.getvalue(), len(rows)


def opcoes_pdf_bytes(ticker, tipo=None, vencimento=None, data_inicio=None, data_fim=None):
    rows = consultar_opcoes(ticker, tipo, vencimento, data_inicio=data_inicio, data_fim=data_fim)
    if not rows:
        return None, 0
    from reportlab.lib import colors

    ticker_limpo = ticker.replace(".SA", "")
    cabecalho = ["Código", "Tipo", "Strike", "Vencimento", "Último", "Variação %",
                 "Vol. Financeiro", "Delta", "Gamma", "Theta", "Vega", "IQ"]
    dados_linhas = []
    for r in rows:
        d = dict(r)
        linha = [
            d.get("codigo") or "—",
            d.get("tipo") or "—",
            _fmt(d.get("strike")),
            str(d.get("vencimento"))[:10] if d.get("vencimento") else "—",
            _fmt(d.get("ultimo")),
            _fmt_pct(d.get("variacao_pct")),
            _fmt_m(d.get("vol_financeiro")),
            _fmt(d.get("delta"), casas=4), _fmt(d.get("gamma"), casas=4),
            _fmt(d.get("theta"), casas=4), _fmt(d.get("vega"),  casas=4),
            _fmt(d.get("iq_calc")),
        ]
        variacao = d.get("variacao_pct")
        cor = None
        if variacao is not None:
            cor = colors.HexColor("#1A7A3C") if variacao >= 0 else colors.HexColor("#B3261E")
        dados_linhas.append((linha, cor))

    subtitulo = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}  ·  {len(rows):,} contratos"
    if tipo:
        subtitulo += f"  ·  Tipo: {tipo}"
    if vencimento:
        subtitulo += f"  ·  Vencimento: {vencimento}"

    buf = io.BytesIO()
    _tabela_pdf(
        buf,
        titulo=f"Cadeia de Opções — {ticker_limpo}",
        subtitulo=subtitulo,
        cabecalho=cabecalho, dados_linhas=dados_linhas, col_cor=5,
        fonte_tamanho=7.5,
    )
    return buf.getvalue(), len(rows)


# ─── Dispatch por nome de formato — usado por api.py ───────────────────────

COTACOES_BYTES = {"csv": cotacoes_csv_bytes, "xlsx": cotacoes_xlsx_bytes, "pdf": cotacoes_pdf_bytes}
OPCOES_BYTES   = {"csv": opcoes_csv_bytes,   "xlsx": opcoes_xlsx_bytes,   "pdf": opcoes_pdf_bytes}
