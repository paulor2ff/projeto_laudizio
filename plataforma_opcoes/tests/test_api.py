"""
Testes para api.py — autenticação Bearer, endpoints REST, rate limit e
WebSocket. api.py não tinha nenhum teste dedicado antes desta suíte
(0% de cobertura), apesar de ser toda a camada REST/WS da plataforma.

Usa o fixture api_client (ver conftest.py): banco temporário isolado,
scheduler neutralizado (não sobe threads reais) e token conhecido.
"""

import pytest


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


class TestRaiz:
    def test_raiz_serve_dashboard(self, api_client):
        client, _ = api_client
        r = client.get("/")
        assert r.status_code == 200
        assert "text/html" in r.headers["content-type"]

    def test_docs_disponivel(self, api_client):
        client, _ = api_client
        r = client.get("/docs")
        assert r.status_code == 200


class TestAutenticacao:
    def test_sem_token_retorna_401(self, api_client):
        client, _ = api_client
        r = client.get("/tickers")
        assert r.status_code == 401

    def test_token_errado_retorna_401(self, api_client):
        client, _ = api_client
        r = client.get("/tickers", headers=_auth("token-errado"))
        assert r.status_code == 401

    def test_token_correto_retorna_200(self, api_client):
        client, token = api_client
        r = client.get("/tickers", headers=_auth(token))
        assert r.status_code == 200

    def test_sem_nenhum_token_configurado_modo_aberto(self, monkeypatch, tmp_path):
        """
        Quando _obter_token_activo() devolve "" (nem API_TOKEN nem ficheiro
        de token), verificar_token() cai em 'modo aberto' (return True).

        Testado chamando verificar_token() diretamente como unidade, em vez
        de subir a app completa via TestClient: passando pelo lifespan real,
        inicializar_token() já teria gerado e gravado um token novo antes de
        qualquer pedido chegar, tornando este ramo inatingível nesse caminho
        — o que é, em si, uma observação útil (ver relatório).
        """
        import config
        monkeypatch.setattr(config, "API_TOKEN", "")

        import api
        monkeypatch.setattr(api, "_TOKEN_FILE", str(tmp_path / "nao_existe.txt"))

        assert api.verificar_token(credentials=None) is True


class TestTickers:
    def test_lista_tickers(self, api_client):
        client, token = api_client
        r = client.get("/tickers", headers=_auth(token))
        assert r.status_code == 200
        body = r.json()
        assert "tickers" in body and "total" in body
        assert body["total"] == len(body["tickers"])


class TestResumo:
    def test_estrutura_do_resumo(self, api_client):
        client, token = api_client
        r = client.get("/resumo", headers=_auth(token))
        assert r.status_code == 200
        for chave in ["cotacoes", "opcoes", "tickers"]:
            assert chave in r.json()


class TestStatus:
    def test_estrutura_do_status(self, api_client):
        client, token = api_client
        r = client.get("/status", headers=_auth(token))
        assert r.status_code == 200
        body = r.json()
        for chave in ["versao", "timestamp", "ws_clientes", "scheduler", "auth_configurada"]:
            assert chave in body
        assert body["scheduler"]["rodando"] is False  # neutralizado pelo fixture


class TestCotacaoAtual:
    def test_preco_atual(self, api_client, yf_ticker):
        client, token = api_client
        yf_ticker.fast_info.last_price = 21.0
        yf_ticker.fast_info.previous_close = 20.0
        r = client.get("/cotacao-atual/BBAS3", headers=_auth(token))
        assert r.status_code == 200
        body = r.json()
        assert body["ticker"] == "BBAS3.SA"
        assert body["preco"] == 21.0


class TestCotacoesESnapshotsEOpcoes:
    def test_cotacoes_vazio_quando_sem_dados(self, api_client):
        client, token = api_client
        r = client.get("/cotacoes/BBAS3", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["total"] == 0

    def test_cotacoes_reflete_dados_inseridos(self, api_client):
        client, token = api_client
        import database
        database.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)
        ])
        r = client.get("/cotacoes/BBAS3", headers=_auth(token))
        assert r.json()["total"] == 1

    def test_ticker_minusculo_e_normalizado_para_maiuscula(self, api_client):
        """
        GET /cotacoes/bbas3 (minúsculo, sem .SA) deve encontrar os dados
        gravados como 'BBAS3.SA'. Regressão do bug corrigido nesta revisão:
        antes, apenas o sufixo '.SA' era acrescentado sem normalizar a
        caixa, e a comparação no SQLite é sensível a maiúsculas — um
        ticker em minúsculas sempre devolvia 0 resultados.
        """
        client, token = api_client
        import database
        database.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)
        ])
        r = client.get("/cotacoes/bbas3", headers=_auth(token))
        assert r.json()["ticker"] == "BBAS3.SA"
        assert r.json()["total"] == 1

    def test_snapshots_vazio(self, api_client):
        client, token = api_client
        r = client.get("/snapshots/BBAS3", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["snapshots"] == []

    def test_opcoes_filtra_por_tipo(self, api_client):
        client, token = api_client
        import database
        database.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        r = client.get("/opcoes/BBAS3?tipo=PUT", headers=_auth(token))
        assert r.json()["total"] == 0
        r2 = client.get("/opcoes/BBAS3?tipo=CALL", headers=_auth(token))
        assert r2.json()["total"] == 1


class TestGreeksHistorico:
    def test_vazio_quando_sem_dados(self, api_client):
        client, token = api_client
        r = client.get("/greeks-historico/BBAS3", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["total"] == 0

    def test_reflete_historico_salvo(self, api_client):
        client, token = api_client
        import database
        database.salvar_greeks_historico([{
            "opcao_codigo": "T200", "ticker_ativo": "BBAS3.SA",
            "data_hora": "2026-06-17 14:00:00", "delta": 0.5, "gamma": 0.01,
            "theta": -0.5, "vega": 0.1, "rho": 0.05, "otm_atm_itm": "ATM",
            "dist_strike": 0.0, "iq_calc": 50.0, "modelo_usado": "binomial",
            "preco_ativo": 20.0, "taxa_cdi": 0.1075, "div_yield": 0.0,
        }])
        r = client.get("/greeks-historico/BBAS3?codigo=T200", headers=_auth(token))
        assert r.json()["total"] == 1


class TestAlertas:
    def test_lista_vazia_inicialmente(self, api_client, alertas_temp):
        client, token = api_client
        r = client.get("/alertas", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["alertas"] == []

    def test_criar_e_listar_alerta(self, api_client, alertas_temp):
        client, token = api_client
        r = client.post(
            "/alertas",
            params={"ticker": "BBAS3", "tipo": "preco", "operador": ">", "valor": 25.0},
            headers=_auth(token),
        )
        assert r.status_code == 200
        r2 = client.get("/alertas", headers=_auth(token))
        assert r2.json()["total"] == 1

    def test_criar_alerta_invalido_retorna_400(self, api_client, alertas_temp):
        client, token = api_client
        r = client.post(
            "/alertas",
            params={"ticker": "BBAS3", "tipo": "tipo_invalido", "operador": ">", "valor": 25.0},
            headers=_auth(token),
        )
        assert r.status_code == 400

    def test_remover_alerta_existente(self, api_client, alertas_temp):
        client, token = api_client
        criado = client.post(
            "/alertas",
            params={"ticker": "BBAS3", "tipo": "preco", "operador": ">", "valor": 25.0},
            headers=_auth(token),
        ).json()
        r = client.delete(f"/alertas/{criado['id']}", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["removido"] is True

    def test_remover_alerta_inexistente_retorna_404(self, api_client, alertas_temp):
        client, token = api_client
        r = client.delete("/alertas/99999", headers=_auth(token))
        assert r.status_code == 404


class TestColetar:
    def test_coletar_dispara_historico_opcoes_e_greeks(self, api_client, yf_ticker, monkeypatch):
        client, token = api_client
        monkeypatch.setattr("collector.RETRY_DELAY_SEG", 0)
        import pandas as pd
        yf_ticker.history.return_value = pd.DataFrame({
            "Open": [20.0], "High": [20.5], "Low": [19.8],
            "Close": [20.3], "Adj Close": [20.3], "Volume": [1000],
        }, index=pd.to_datetime(["2026-01-02"]))
        yf_ticker.options = ()
        yf_ticker.fast_info.last_price = 20.3
        yf_ticker.fast_info.previous_close = 20.0

        r = client.post("/coletar/BBAS3", headers=_auth(token))
        assert r.status_code == 200
        body = r.json()
        assert body["historico"] == 1
        assert body["opcoes"] == 0  # sem vencimentos no mock

    def test_coletar_todos(self, api_client, monkeypatch):
        client, token = api_client
        monkeypatch.setattr("collector.coletar_historico_todos", lambda: {"BBAS3.SA": 1})
        monkeypatch.setattr("collector.coletar_opcoes_todos", lambda: {"BBAS3.SA": 0})
        monkeypatch.setattr("collector.calcular_greeks_todos", lambda: {"BBAS3.SA": 0})
        r = client.post("/coletar-todos", headers=_auth(token))
        assert r.status_code == 200
        assert r.json()["historico"] == {"BBAS3.SA": 1}


class TestRateLimit:
    def test_sexta_requisicao_seguida_e_bloqueada(self, api_client, yf_ticker, monkeypatch):
        """_RATE_LIMIT_REQ = 5 por _RATE_LIMIT_JAN segundos — a 6ª deve levar 429."""
        client, token = api_client
        monkeypatch.setattr("collector.RETRY_DELAY_SEG", 0)
        yf_ticker.history.return_value = __import__("pandas").DataFrame()  # vazio -> coleta rápida
        yf_ticker.options = ()
        yf_ticker.fast_info.last_price = 20.0
        yf_ticker.fast_info.previous_close = 20.0

        respostas = [
            client.post("/coletar/BBAS3", headers=_auth(token)).status_code
            for _ in range(6)
        ]
        assert respostas[:5] == [200] * 5
        assert respostas[5] == 429


class TestExportarEndpoints:
    def test_cotacoes_sem_dados_retorna_404(self, api_client):
        client, token = api_client
        r = client.get("/exportar/cotacoes/BBAS3", headers=_auth(token))
        assert r.status_code == 404

    def test_cotacoes_xlsx_devolve_arquivo_com_headers_corretos(self, api_client):
        client, token = api_client
        import database
        database.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)
        ])
        r = client.get("/exportar/cotacoes/BBAS3?formato=xlsx", headers=_auth(token))
        assert r.status_code == 200
        assert r.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "cotacoes_BBAS3_" in r.headers["content-disposition"]
        assert "attachment" in r.headers["content-disposition"]

        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(r.content))
        assert "BBAS3" in wb.active["A1"].value

    def test_cotacoes_pdf_devolve_pdf_valido(self, api_client):
        client, token = api_client
        import database
        database.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)
        ])
        r = client.get("/exportar/cotacoes/BBAS3?formato=pdf", headers=_auth(token))
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"

        from pypdf import PdfReader
        import io
        texto = PdfReader(io.BytesIO(r.content)).pages[0].extract_text()
        assert "BBAS3" in texto

    def test_cotacoes_csv_e_o_padrao_documentado_mas_xlsx_e_o_default_real(self, api_client):
        """formato tem default 'xlsx' no endpoint — diferente do default 'csv' da CLI,
        de propósito: quem clica um botão no dashboard quer o arquivo bonito."""
        client, token = api_client
        import database
        database.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)
        ])
        r = client.get("/exportar/cotacoes/BBAS3", headers=_auth(token))
        assert r.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_formato_invalido_retorna_422(self, api_client):
        client, token = api_client
        r = client.get("/exportar/cotacoes/BBAS3?formato=docx", headers=_auth(token))
        assert r.status_code == 422

    def test_sem_token_retorna_401(self, api_client):
        client, _token = api_client
        r = client.get("/exportar/cotacoes/BBAS3")
        assert r.status_code == 401

    def test_opcoes_sem_dados_retorna_404(self, api_client):
        client, token = api_client
        r = client.get("/exportar/opcoes/BBAS3", headers=_auth(token))
        assert r.status_code == 404

    def test_opcoes_xlsx_devolve_arquivo(self, api_client):
        client, token = api_client
        import database
        database.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        r = client.get("/exportar/opcoes/BBAS3?formato=xlsx&tipo=CALL", headers=_auth(token))
        assert r.status_code == 200
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.active.cell(row=5, column=1).value == "T200"

    def test_opcoes_filtro_tipo_sem_resultado_retorna_404(self, api_client):
        client, token = api_client
        import database
        database.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        r = client.get("/exportar/opcoes/BBAS3?tipo=PUT", headers=_auth(token))
        assert r.status_code == 404


class TestWebSocket:
    def test_conexao_recebe_atualizacao(self, api_client, yf_ticker):
        client, token = api_client
        yf_ticker.fast_info.last_price = 20.0
        yf_ticker.fast_info.previous_close = 20.0
        with client.websocket_connect(f"/ws/BBAS3?token={token}") as ws:
            msg = ws.receive_json()
            assert msg["tipo"] == "update"
            assert msg["ticker"] == "BBAS3.SA"
            assert "cotacao" in msg and "opcoes" in msg

    def test_token_invalido_fecha_conexao(self, api_client):
        from starlette.websockets import WebSocketDisconnect
        client, token = api_client
        with pytest.raises(WebSocketDisconnect):
            with client.websocket_connect("/ws/BBAS3?token=errado") as ws:
                ws.receive_json()
