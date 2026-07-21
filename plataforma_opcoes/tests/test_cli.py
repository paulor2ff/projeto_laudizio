"""
Testes para cli.py — formatação, validação, impressão de tabelas,
exportação CSV e o dispatch de main() para cada comando. cli.py não
tinha nenhum teste dedicado antes desta suíte (0% de cobertura), apesar
de ser o maior módulo do projecto (366 linhas).

Usa o fixture cli_env (ver conftest.py): banco temporário isolado,
LOG_PATH isolado e cwd temporário (para não sujar o projecto real com
CSVs/logs de teste).
"""

import csv
import sys
import pytest

import cli


# ─── Formatação ─────────────────────────────────────────────────────────────

class TestFmt:
    def test_none_usa_default(self):
        assert cli._fmt(None) == "—"

    def test_valor_valido(self):
        assert cli._fmt(20.12345, casas=2) == "20.12"

    def test_valor_invalido_cai_para_str(self):
        assert cli._fmt("abc") == "abc"


class TestFmtI:
    def test_none_usa_default(self):
        assert cli._fmt_i(None) == "—"

    def test_milhar_com_ponto_br(self):
        assert cli._fmt_i(1_500_000) == "1.500.000"


class TestFmtPct:
    def test_none_usa_default(self):
        assert cli._fmt_pct(None) == "—"

    def test_positivo_tem_sinal(self):
        assert cli._fmt_pct(3.2) == "+3.20%"

    def test_negativo_tem_sinal(self):
        assert cli._fmt_pct(-1.5) == "-1.50%"


class TestFmtM:
    def test_none_usa_default(self):
        assert cli._fmt_m(None) == "—"

    def test_abaixo_de_mil(self):
        assert cli._fmt_m(500) == "R$ 500.00"

    def test_milhares_em_k(self):
        assert cli._fmt_m(150_000) == "R$ 150.0K"

    def test_milhoes_em_m(self):
        assert cli._fmt_m(2_500_000) == "R$ 2.5M"


class TestValidarData:
    def test_data_valida(self):
        assert cli._validar_data("2026-06-17", "--de") == "2026-06-17"

    def test_formato_sem_zeros_sai_com_erro(self):
        with pytest.raises(SystemExit):
            cli._validar_data("2026-6-7", "--de")

    def test_data_inexistente_sai_com_erro(self):
        with pytest.raises(SystemExit):
            cli._validar_data("2026-02-30", "--de")  # 30 de fevereiro não existe

    def test_texto_arbitrario_sai_com_erro(self):
        with pytest.raises(SystemExit):
            cli._validar_data("ontem", "--de")


class TestNormalizarTicker:
    def test_maiuscula_sem_sufixo(self):
        assert cli._normalizar_ticker("BBAS3") == "BBAS3.SA"

    def test_minuscula_e_normalizada(self):
        """Regressão do mesmo bug corrigido em api.py: minúsculas viravam
        sufixo sem normalizar a caixa, o que não batia com o dado gravado."""
        assert cli._normalizar_ticker("bbas3") == "BBAS3.SA"

    def test_ja_com_sufixo_nao_duplica(self):
        assert cli._normalizar_ticker("BBAS3.SA") == "BBAS3.SA"

    def test_sufixo_minusculo(self):
        assert cli._normalizar_ticker("bbas3.sa") == "BBAS3.SA"


# ─── Impressão (capsys) ─────────────────────────────────────────────────────

class TestImprimirCotacoes:
    def test_lista_vazia(self, capsys):
        cli.imprimir_cotacoes([])
        assert "Nenhum registro" in capsys.readouterr().out

    def test_lista_populada(self, capsys):
        rows = [{"data": "2026-06-17", "abertura": 20.0, "maxima": 20.5,
                  "minima": 19.8, "fechamento": 20.3, "volume": 1_000_000}]
        cli.imprimir_cotacoes(rows, titulo="Teste")
        out = capsys.readouterr().out
        assert "Teste" in out
        assert "2026-06-17" in out
        assert "Total: 1 registros" in out


class TestImprimirOpcoes:
    def test_lista_vazia(self, capsys):
        cli.imprimir_opcoes([])
        assert "Nenhuma opção" in capsys.readouterr().out

    def test_lista_populada(self, capsys):
        rows = [{"codigo": "BBAS3C200", "tipo": "CALL", "otm_atm_itm": "ATM",
                  "strike": 20.0, "vencimento": "2027-06-18", "ultimo": 1.5,
                  "variacao_pct": 3.2, "num_negocios": 150, "vol_financeiro": 225_000.0,
                  "delta": 0.5, "gamma": 0.01, "theta": -0.5, "vega": 0.1, "iq_calc": 50.0}]
        cli.imprimir_opcoes(rows)
        out = capsys.readouterr().out
        assert "BBAS3C200" in out
        assert "↑CALL" in out


class TestImprimirResumo:
    def test_banco_vazio(self, db_temp, capsys):
        cli.imprimir_resumo()
        out = capsys.readouterr().out
        assert "Resumo da Plataforma" in out
        assert "Tickers monitorados : 0" in out

    def test_banco_com_dados(self, db_temp, capsys):
        db_temp.upsert_cotacoes([("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)])
        cli.imprimir_resumo()
        out = capsys.readouterr().out
        assert "BBAS3" in out


class TestImprimirHistoricoGreeks:
    def test_lista_vazia(self, capsys):
        cli.imprimir_historico_greeks([])
        assert "Nenhum registro de histórico" in capsys.readouterr().out

    def test_lista_populada(self, capsys):
        rows = [{"data_hora": "2026-06-17 14:00:00", "opcao_codigo": "BBAS3C200",
                  "delta": 0.5, "gamma": 0.01, "theta": -0.5, "vega": 0.1,
                  "iq_calc": 50.0, "modelo_usado": "binomial"}]
        cli.imprimir_historico_greeks(rows)
        assert "BBAS3C200" in capsys.readouterr().out


class TestImprimirAlertas:
    def test_lista_vazia(self, capsys):
        cli.imprimir_alertas([])
        assert "Nenhum alerta" in capsys.readouterr().out

    def test_lista_populada(self, capsys):
        alertas = [{"id": 1, "ticker": "BBAS3.SA", "tipo": "preco", "codigo": None,
                     "operador": ">", "valor": 25.0, "activo": True, "disparado": False,
                     "cooldown_min": 15, "ultimo_val": 20.5}]
        cli.imprimir_alertas(alertas)
        out = capsys.readouterr().out
        assert "BBAS3" in out
        assert "✅" in out


# ─── Exportação CSV ─────────────────────────────────────────────────────────

class TestExportarCotacoesCSV:
    def test_sem_dados(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        cli.exportar_cotacoes_csv("BBAS3.SA")
        assert "Nenhum dado" in capsys.readouterr().out
        assert list(tmp_path.iterdir()) == []

    def test_com_dados_gera_csv_valido(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_cotacoes([("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)])
        cli.exportar_cotacoes_csv("BBAS3.SA")
        arquivos = list(tmp_path.glob("cotacoes_BBAS3_*.csv"))
        assert len(arquivos) == 1
        with open(arquivos[0], encoding="utf-8-sig") as f:
            linhas = list(csv.DictReader(f))
        assert len(linhas) == 1
        assert linhas[0]["ticker"] == "BBAS3.SA"


class TestExportarOpcoesCSV:
    def test_com_dados_gera_csv_valido(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        cli.exportar_opcoes_csv("BBAS3.SA")
        arquivos = list(tmp_path.glob("opcoes_BBAS3_*.csv"))
        assert len(arquivos) == 1
        with open(arquivos[0], encoding="utf-8-sig") as f:
            linhas = list(csv.DictReader(f))
        assert linhas[0]["codigo"] == "T200"


# ─── main() — dispatch de comandos ──────────────────────────────────────────

def _rodar(monkeypatch, cli_env, *args):
    monkeypatch.setattr(sys, "argv", ["cli.py", *args])
    cli_env.main()


class TestExportarCotacoesXlsx:
    def test_sem_dados(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        cli.exportar_cotacoes_xlsx("BBAS3.SA")
        assert "Nenhum dado" in capsys.readouterr().out
        assert list(tmp_path.iterdir()) == []

    def test_com_dados_gera_xlsx_valido(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000),  # fechou em alta
            ("BBAS3.SA", "2026-01-05", 20.3, 20.4, 19.0, 19.5, 19.5, 2000),  # fechou em baixa
        ])
        cli.exportar_cotacoes_xlsx("BBAS3.SA")
        arquivos = list(tmp_path.glob("cotacoes_BBAS3_*.xlsx"))
        assert len(arquivos) == 1

        from openpyxl import load_workbook
        wb = load_workbook(arquivos[0])
        ws = wb.active
        assert "BBAS3" in ws["A1"].value
        cabecalhos = [ws.cell(row=4, column=c).value for c in range(1, 9)]
        assert cabecalhos == ["Ticker", "Data", "Abertura", "Máxima", "Mínima",
                               "Fechamento", "Fech. Ajustado", "Volume"]
        # 2 linhas de dados nas linhas 5 e 6 (cabeçalho na linha 4)
        # consultar_cotacoes devolve mais recente primeiro
        assert str(ws.cell(row=5, column=2).value)[:10] == "2026-01-05"
        assert str(ws.cell(row=6, column=2).value)[:10] == "2026-01-02"
        # a mais recente (05/01) fechou em baixa -> vermelho; a mais antiga (02/01) em alta -> verde
        assert ws.cell(row=5, column=6).font.color.rgb.endswith("B3261E")
        assert ws.cell(row=6, column=6).font.color.rgb.endswith("1A7A3C")


class TestExportarOpcoesXlsx:
    def _inserir(self, db_temp, variacao=1.0):
        db_temp.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": variacao, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])

    def test_sem_dados(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        cli.exportar_opcoes_xlsx("BBAS3.SA")
        assert "Nenhuma opção" in capsys.readouterr().out

    def test_com_dados_gera_xlsx_valido(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        self._inserir(db_temp, variacao=3.2)
        cli.exportar_opcoes_xlsx("BBAS3.SA")
        arquivos = list(tmp_path.glob("opcoes_BBAS3_*.xlsx"))
        assert len(arquivos) == 1

        from openpyxl import load_workbook
        wb = load_workbook(arquivos[0])
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "Código"
        assert ws.cell(row=5, column=1).value == "T200"
        # variação positiva -> coluna "Variação %" (7ª) em verde
        assert ws.cell(row=5, column=7).font.color.rgb.endswith("1A7A3C")

    def test_variacao_negativa_fica_vermelha(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        self._inserir(db_temp, variacao=-2.5)
        cli.exportar_opcoes_xlsx("BBAS3.SA")
        from openpyxl import load_workbook
        wb = load_workbook(next(tmp_path.glob("opcoes_BBAS3_*.xlsx")))
        ws = wb.active
        assert ws.cell(row=5, column=7).font.color.rgb.endswith("B3261E")


class TestExportarCotacoesPdf:
    def test_sem_dados(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        cli.exportar_cotacoes_pdf("BBAS3.SA")
        assert "Nenhum dado" in capsys.readouterr().out
        assert list(tmp_path.iterdir()) == []

    def test_com_dados_gera_pdf_com_conteudo_esperado(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_cotacoes([("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)])
        cli.exportar_cotacoes_pdf("BBAS3.SA")
        arquivos = list(tmp_path.glob("cotacoes_BBAS3_*.pdf"))
        assert len(arquivos) == 1

        from pypdf import PdfReader
        texto = PdfReader(str(arquivos[0])).pages[0].extract_text()
        assert "BBAS3" in texto
        assert "2026-01-02" in texto
        assert "20.30" in texto  # fechamento formatado por _fmt


class TestExportarOpcoesPdf:
    def test_com_dados_gera_pdf_com_conteudo_esperado(self, db_temp, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 3.2, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        cli.exportar_opcoes_pdf("BBAS3.SA", tipo="CALL")
        arquivos = list(tmp_path.glob("opcoes_BBAS3_*.pdf"))
        assert len(arquivos) == 1

        from pypdf import PdfReader
        texto = PdfReader(str(arquivos[0])).pages[0].extract_text()
        assert "T200" in texto
        assert "Tipo: CALL" in texto  # veio do subtítulo, já que --tipo foi passado


class TestMainFormatoExportacao:
    def test_exportar_com_formato_xlsx(self, monkeypatch, cli_env, db_temp, tmp_path, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_cotacoes([("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)])
        _rodar(monkeypatch, cli_env, "--exportar", "--ticker", "BBAS3", "--formato", "xlsx")
        assert list(tmp_path.glob("cotacoes_BBAS3_*.xlsx"))
        assert not list(tmp_path.glob("cotacoes_BBAS3_*.csv"))

    def test_exportar_opcoes_com_formato_pdf(self, monkeypatch, cli_env, db_temp, tmp_path, capsys):
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_opcoes([{
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }])
        _rodar(monkeypatch, cli_env, "--exportar-opcoes", "--ticker", "BBAS3", "--formato", "pdf")
        assert list(tmp_path.glob("opcoes_BBAS3_*.pdf"))

    def test_formato_invalido_sai_com_erro(self, monkeypatch, cli_env):
        monkeypatch.setattr(sys, "argv",
                             ["cli.py", "--exportar", "--formato", "docx"])
        with pytest.raises(SystemExit):
            cli_env.main()

    def test_formato_padrao_continua_csv(self, monkeypatch, cli_env, db_temp, tmp_path, capsys):
        """Sem --formato explícito, o comportamento de sempre (csv) não muda."""
        monkeypatch.chdir(tmp_path)
        db_temp.upsert_cotacoes([("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000)])
        _rodar(monkeypatch, cli_env, "--exportar", "--ticker", "BBAS3")
        assert list(tmp_path.glob("cotacoes_BBAS3_*.csv"))


class TestMainSemArgumentos:
    def test_sem_flags_imprime_ajuda(self, monkeypatch, cli_env, capsys):
        _rodar(monkeypatch, cli_env)
        out = capsys.readouterr().out
        assert "usage" in out.lower() or "uso" in out.lower()


class TestMainConsultas:
    def test_resumo(self, monkeypatch, cli_env, capsys):
        _rodar(monkeypatch, cli_env, "--resumo")
        assert "Resumo da Plataforma" in capsys.readouterr().out

    def test_consultar_sem_dados(self, monkeypatch, cli_env, capsys):
        _rodar(monkeypatch, cli_env, "--consultar", "--ticker", "BBAS3")
        assert "Nenhum registro" in capsys.readouterr().out

    def test_consultar_com_data_invalida_sai_com_erro(self, monkeypatch, cli_env):
        monkeypatch.setattr(sys, "argv",
                             ["cli.py", "--consultar", "--de", "2026-2-30"])
        with pytest.raises(SystemExit):
            cli_env.main()

    def test_opcoes_com_tipo_invalido_sai_com_erro(self, monkeypatch, cli_env):
        monkeypatch.setattr(sys, "argv",
                             ["cli.py", "--opcoes", "--tipo", "XPTO"])
        with pytest.raises(SystemExit):
            cli_env.main()

    def test_status(self, monkeypatch, cli_env, capsys):
        _rodar(monkeypatch, cli_env, "--status")
        assert "Status da Plataforma" in capsys.readouterr().out


class TestMainColeta:
    def test_coletar_chama_collector(self, monkeypatch, cli_env, capsys):
        chamadas = {}
        monkeypatch.setattr(
            "collector.coletar_historico",
            lambda ticker, periodo: chamadas.setdefault("args", (ticker, periodo)) or 3
        )
        _rodar(monkeypatch, cli_env, "--coletar", "--ticker", "BBAS3", "--periodo", "5d")
        assert chamadas["args"] == ("BBAS3.SA", "5d")

    def test_coletar_todos_imprime_total(self, monkeypatch, cli_env, capsys):
        monkeypatch.setattr(
            "collector.coletar_historico_todos",
            lambda periodo: {"BBAS3.SA": 3, "PETR4.SA": 2}
        )
        _rodar(monkeypatch, cli_env, "--coletar-todos")
        out = capsys.readouterr().out
        assert "5" in out and "2 tickers" in out

    def test_coletar_opcoes_todos_imprime_total(self, monkeypatch, cli_env, capsys):
        monkeypatch.setattr(
            "collector.coletar_opcoes_todos",
            lambda: {"BBAS3.SA": 4, "PETR4.SA": 6}
        )
        _rodar(monkeypatch, cli_env, "--coletar-opcoes-todos")
        out = capsys.readouterr().out
        assert "10" in out and "2 tickers" in out

    def test_calcular_greeks_todos_imprime_total(self, monkeypatch, cli_env, capsys):
        monkeypatch.setattr(
            "collector.calcular_greeks_todos",
            lambda: {"BBAS3.SA": 4, "PETR4.SA": 6}
        )
        _rodar(monkeypatch, cli_env, "--calcular-greeks-todos")
        out = capsys.readouterr().out
        assert "10" in out and "2 tickers" in out


class TestMainAlertas:
    def test_alerta_add_e_listar(self, monkeypatch, cli_env, alertas_temp, capsys):
        _rodar(monkeypatch, cli_env, "--alerta-add", "--ticker", "BBAS3",
               "--tipo-alerta", "preco", "--operador", ">", "--valor", "25")
        assert "criado" in capsys.readouterr().out.lower()

        _rodar(monkeypatch, cli_env, "--alertas")
        assert "BBAS3" in capsys.readouterr().out

    def test_alerta_add_sem_campos_obrigatorios_sai_com_erro(self, monkeypatch, cli_env, alertas_temp):
        monkeypatch.setattr(sys, "argv", ["cli.py", "--alerta-add", "--ticker", "BBAS3"])
        with pytest.raises(SystemExit):
            cli_env.main()

    def test_alerta_remover_inexistente_sai_com_erro(self, monkeypatch, cli_env, alertas_temp):
        monkeypatch.setattr(sys, "argv", ["cli.py", "--alerta-remover", "--id-alerta", "999"])
        with pytest.raises(SystemExit):
            cli_env.main()


class TestMainLicenca:
    def test_licenca_sem_instalacao(self, monkeypatch, cli_env, licenca_temp, capsys):
        _rodar(monkeypatch, cli_env, "--licenca")
        out = capsys.readouterr().out
        assert "sem_licenca" in out

    def test_licenca_importar(self, monkeypatch, cli_env, licenca_temp, tmp_path, capsys):
        import json
        dados = licenca_temp.emitir(dias_validade=10)
        caminho = tmp_path / "recebida.json"
        caminho.write_text(json.dumps(dados))
        _rodar(monkeypatch, cli_env, "--licenca-importar", str(caminho))
        assert "importada com sucesso" in capsys.readouterr().out.lower()

    def test_licenca_importar_arquivo_inexistente_sai_com_erro(self, monkeypatch, cli_env, licenca_temp):
        monkeypatch.setattr(sys, "argv",
                             ["cli.py", "--licenca-importar", "/tmp/nao_existe_xyz.json"])
        with pytest.raises(SystemExit):
            cli_env.main()


class TestMainDashboard:
    def test_dashboard_chama_uvicorn_run_sem_bloquear(self, monkeypatch, cli_env, capsys):
        import uvicorn
        chamadas = {}
        monkeypatch.setattr(
            uvicorn, "run",
            lambda app, **kwargs: chamadas.setdefault("kwargs", kwargs)
        )
        _rodar(monkeypatch, cli_env, "--dashboard")
        assert "kwargs" in chamadas
        assert chamadas["kwargs"]["workers"] == 1
