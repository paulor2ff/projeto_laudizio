"""
Testes para exportadores.py — o módulo partilhado entre cli.py (grava em
disco) e api.py (devolve por HTTP). A lógica detalhada de estilo (cores,
zebra, cabeçalhos) já é exercitada via tests/test_cli.py, que chama estas
mesmas funções por baixo — aqui o foco é o contrato em bytes: devolve
(None, 0) sem dados, (bytes, contagem) com dados, e o conteúdo é
decodificável/parseável pelo formato certo.
"""

import csv
import io

import pytest

import exportadores as exp


class TestNomeExportacao:
    def test_formato_padrao(self):
        nome = exp._nome_exportacao("cotacoes", "BBAS3.SA", "xlsx")
        assert nome.startswith("cotacoes_BBAS3_")
        assert nome.endswith(".xlsx")


class TestCotacoesBytes:
    def _inserir(self, db_temp):
        db_temp.upsert_cotacoes([
            ("BBAS3.SA", "2026-01-02", 20.0, 20.5, 19.8, 20.3, 20.3, 1000),
            ("BBAS3.SA", "2026-01-05", 20.3, 20.4, 19.0, 19.5, 19.5, 2000),
        ])

    @pytest.mark.parametrize("gerador", [
        exp.cotacoes_csv_bytes, exp.cotacoes_xlsx_bytes, exp.cotacoes_pdf_bytes,
    ])
    def test_sem_dados_devolve_none_e_zero(self, db_temp, gerador):
        conteudo, contagem = gerador("BBAS3.SA")
        assert conteudo is None
        assert contagem == 0

    def test_csv_bytes_decodifica_e_tem_bom(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.cotacoes_csv_bytes("BBAS3.SA")
        assert contagem == 2
        assert conteudo.startswith("\ufeff".encode("utf-8"))
        texto = conteudo.decode("utf-8-sig")
        linhas = list(csv.DictReader(io.StringIO(texto)))
        assert len(linhas) == 2
        assert linhas[0]["ticker"] == "BBAS3.SA"

    def test_xlsx_bytes_e_uma_planilha_valida(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.cotacoes_xlsx_bytes("BBAS3.SA")
        assert contagem == 2
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(conteudo))
        ws = wb.active
        assert "BBAS3" in ws["A1"].value

    def test_pdf_bytes_e_um_pdf_valido_com_conteudo(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.cotacoes_pdf_bytes("BBAS3.SA")
        assert contagem == 2
        from pypdf import PdfReader
        texto = PdfReader(io.BytesIO(conteudo)).pages[0].extract_text()
        assert "BBAS3" in texto


class TestOpcoesBytes:
    def _inserir(self, db_temp, **overrides):
        registro = {
            "ticker_ativo": "BBAS3.SA", "codigo": "T200", "tipo": "CALL",
            "modelo": None, "strike": 20.0, "vencimento": "2027-06-18",
            "ultimo": 1.5, "variacao_pct": 1.0, "data_hora": "2026-06-17 14:00:00",
            "num_negocios": 100, "vol_financeiro": 150_000.0, "vol_implícita": 0.30,
            "iq": None, "coberto": None, "descoberto": None, "travado": None,
            "titulares": None, "lancadores": None, "fonte": "test",
        }
        registro.update(overrides)
        db_temp.upsert_opcoes([registro])

    @pytest.mark.parametrize("gerador", [
        exp.opcoes_csv_bytes, exp.opcoes_xlsx_bytes, exp.opcoes_pdf_bytes,
    ])
    def test_sem_dados_devolve_none_e_zero(self, db_temp, gerador):
        conteudo, contagem = gerador("BBAS3.SA")
        assert conteudo is None
        assert contagem == 0

    def test_csv_bytes_decodifica(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.opcoes_csv_bytes("BBAS3.SA")
        assert contagem == 1
        texto = conteudo.decode("utf-8-sig")
        linhas = list(csv.DictReader(io.StringIO(texto)))
        assert linhas[0]["codigo"] == "T200"

    def test_xlsx_bytes_e_uma_planilha_valida(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.opcoes_xlsx_bytes("BBAS3.SA")
        assert contagem == 1
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(conteudo))
        ws = wb.active
        assert ws.cell(row=5, column=1).value == "T200"

    def test_pdf_bytes_com_filtro_de_tipo(self, db_temp):
        self._inserir(db_temp)
        conteudo, contagem = exp.opcoes_pdf_bytes("BBAS3.SA", tipo="CALL")
        assert contagem == 1
        from pypdf import PdfReader
        texto = PdfReader(io.BytesIO(conteudo)).pages[0].extract_text()
        assert "T200" in texto

    def test_filtro_por_tipo_exclui_o_que_nao_bate(self, db_temp):
        self._inserir(db_temp, codigo="CALL1", tipo="CALL")
        self._inserir(db_temp, codigo="PUT1", tipo="PUT")
        _conteudo, contagem = exp.opcoes_csv_bytes("BBAS3.SA", tipo="PUT")
        assert contagem == 1


class TestDispatchPorFormato:
    def test_cotacoes_bytes_tem_as_tres_chaves(self):
        assert set(exp.COTACOES_BYTES) == {"csv", "xlsx", "pdf"}

    def test_opcoes_bytes_tem_as_tres_chaves(self):
        assert set(exp.OPCOES_BYTES) == {"csv", "xlsx", "pdf"}

    def test_media_types_tem_as_tres_chaves(self):
        assert set(exp.MEDIA_TYPES) == {"csv", "xlsx", "pdf"}
